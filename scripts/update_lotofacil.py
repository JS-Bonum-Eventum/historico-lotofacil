"""
update_lotofacil.py
Lotofácil: 15 números (Bola1-Bola15), range 1-25
Cabeçalho na linha 1, dados a partir da linha 2, bolas nas colunas C-Q (índice 2-16)
"""

import requests
from io import BytesIO

EXCEL_URL = "https://servicebus2.caixa.gov.br/portaldeloterias/api/resultados/download?modalidade=Lotof%C3%A1cil"
OUTPUT_FILE = "historico.txt"
SEQUENCE_LENGTH = 15
MAX_NUMBER = 25
BALL_COL_START = 2
BALL_COL_END = 17

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0 Safari/537.36",
    "Accept": "*/*",
    "Referer": "https://loterias.caixa.gov.br/Paginas/Lotofacil.aspx",
}

def to_int(val):
    if val is None or val == '':
        return None
    try:
        # Remove espaços, substitui vírgula por ponto
        s = str(val).strip().replace(',', '.').split('.')[0]
        return int(s)
    except (ValueError, TypeError):
        return None

def download_excel():
    print("Baixando Excel da Lotofácil...")
    resp = requests.get(EXCEL_URL, headers=HEADERS, timeout=60)
    resp.raise_for_status()
    content = resp.content
    print(f"Download OK: {len(content)} bytes")
    if content[:2] == b'PK':
        fmt = 'xlsx'
    elif content[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        fmt = 'xls'
    else:
        fmt = 'xlsx'
    print(f"Formato: {fmt}")
    return BytesIO(content), fmt

def parse_xlsx(file_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    ws = wb.active
    print(f"openpyxl: {ws.max_row} linhas x {ws.max_column} colunas")
    rows = [list(row) for row in ws.iter_rows(min_row=1, values_only=True)]
    wb.close()
    return rows

def parse_xls(file_bytes):
    import xlrd
    wb = xlrd.open_workbook(file_contents=file_bytes.read())
    ws = wb.sheet_by_index(0)
    print(f"xlrd: {ws.nrows} linhas x {ws.ncols} colunas")
    return [ws.row_values(i) for i in range(ws.nrows)]

def extract_sequences(rows):
    print(f"Total linhas no arquivo: {len(rows)}")

    skipped_empty = 0
    skipped_invalid = 0
    skipped_range = 0
    sequences = []

    for idx, row in enumerate(rows[1:], start=2):  # pula cabeçalho
        # Pula linha vazia
        if not row or row[0] is None or str(row[0]).strip() == '':
            skipped_empty += 1
            continue

        nums = []
        raw_vals = []
        for i in range(BALL_COL_START, BALL_COL_END):
            val = row[i] if i < len(row) else None
            raw_vals.append(val)
            n = to_int(val)
            if n is not None:
                nums.append(n)

        # Verifica quantidade
        if len(nums) != SEQUENCE_LENGTH:
            skipped_invalid += 1
            if skipped_invalid <= 5:
                print(f"  Linha {idx} inválida ({len(nums)} nums): raw={raw_vals}")
            continue

        # Verifica range — aceita mesmo fora do range e loga
        out_of_range = [n for n in nums if n < 1 or n > MAX_NUMBER]
        if out_of_range:
            skipped_range += 1
            if skipped_range <= 5:
                print(f"  Linha {idx} fora do range: {out_of_range} | nums={nums}")
            continue

        sorted_nums = sorted(set(nums))
        # Se set() removeu duplicatas, ainda tenta usar se tiver 15 únicos
        if len(sorted_nums) == SEQUENCE_LENGTH:
            sequences.append(sorted_nums)
        else:
            skipped_invalid += 1
            if skipped_invalid <= 5:
                print(f"  Linha {idx} com duplicatas: {nums}")

    print(f"Sequências válidas: {len(sequences)}")
    print(f"Puladas (vazias): {skipped_empty}")
    print(f"Puladas (inválidas): {skipped_invalid}")
    print(f"Puladas (fora do range): {skipped_range}")
    return sequences

def write_output(sequences):
    seen = set()
    unique = []
    for s in sequences:
        key = ','.join(map(str, s))
        if key not in seen:
            seen.add(key)
            unique.append(s)
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        for seq in unique:
            f.write(' '.join(map(str, seq)) + '\n')
    print(f"historico.txt: {len(unique)} linhas escritas")

def main():
    file_bytes, fmt = download_excel()
    rows = parse_xls(file_bytes) if fmt == 'xls' else parse_xlsx(file_bytes)
    sequences = extract_sequences(rows)
    if not sequences:
        raise ValueError("Nenhuma sequência válida encontrada!")
    write_output(sequences)
    print("✅ Lotofácil atualizado com sucesso!")

if __name__ == "__main__":
    main()
